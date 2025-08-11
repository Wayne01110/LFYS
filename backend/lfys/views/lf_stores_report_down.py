from datetime import datetime

import requests
import openpyxl
from io import BytesIO
from django.http import FileResponse
from rest_framework.views import APIView


class LFStoresReportDown(APIView):
    permission_classes = []

    """
        add by Wayne 2025-07-03  各门店报表汇总下载
    """

    def post(self, request):
        login_infos = [
            {"name": "长春-万象店", "tenantId": "28"},
            {"name": "长春-远大店", "tenantId": "2"},
            {"name": "长春-欧亚店", "tenantId": "7"},
            {"name": "哈尔滨-万象店", "tenantId": "34"},
            {"name": "长春-东樾店", "tenantId": "16"},
            {"name": "郑州-万象店", "tenantId": "35"},
            {"name": "呼和浩特-万象店", "tenantId": "36"},
        ]

        # 登录接口
        login_url = "https://lfyisheng.hospital.realmerit.com.cn/api/v1/login"
        # 各门店客户资产保有统计（含定金、储值、赠送金、积分）
        report_url = "https://lfyisheng.hospital.realmerit.com.cn/api/report/v1/hospitalDepositBalance/report"
        # 咨询接诊记录
        admission_url = "https://lfyisheng.hospital.realmerit.com.cn/api/v1/consultation/page"


        base_login_payload = {
            "account": "18188889999",
            "password": "PGHFbsn1XxiZ+qe0+uQBHSVaGtokdLuYMpoZbb2UUWakOT4FxiO39iBtovzfGi6TpK6kDbkYwaOTRdr2BMbDhMAiyrJwdgtj9sjYy6RYKyLTDpkBPQe0f4sgkejnQ9Go2MDXlQD1rMlwtzlBVIuZYMvx96Zd0Kb7oHoQEuypTBY=",
            "hospitalKey": "lfys4123",
            "type": 1,
            "mac": "",
            "loginType": "PC_WEB"
        }

        action = request.data.get("action", "")
        # 各门店客户资产保有统计（含定金、储值、赠送金、积分）
        if action == "client_asset":
            all_customer_records = []
            for info in login_infos:
                login_payload = base_login_payload.copy()
                login_payload["tenantId"] = info["tenantId"]

                try:
                    login_resp = requests.post(login_url, json=login_payload, headers={"Content-Type": "application/json"})
                    token = login_resp.json().get("data", {}).get("user", {}).get("token")
                    if not token:
                        continue

                    page = 1
                    size = 30000
                    total_pages = 1
                    while page <= total_pages:
                        report_payload = {
                            "balanceFlagList": [],
                            "givenBalanceFlagList": [],
                            "depositFlagList": [],
                            "integralFlagList": [],
                            "index": page,
                            "size": size
                        }

                        report_resp = requests.post(
                            report_url,
                            json=report_payload,
                            headers={"Content-Type": "application/json", "authorization": token}
                        )

                        data = report_resp.json().get("data", {})
                        page_data = data.get("page", {})
                        all_customer_records.extend(page_data.get("records", []))
                        total_pages = page_data.get("pages", 1)
                        page += 1

                except Exception:
                    continue

            # 开始写 Excel 文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "客户资产保有统计"

            # 写第一行标题
            ws.append(["客户资产保有统计"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)

            # 第二行字段头
            headers = [
                "客户", "会员号", "vip等级", "客户类型", "归属咨询",
                "定金累计", "定金退款", "定金抵扣", "定金保有",
                "储值累计", "储值退款", "储值抵扣", "储值保有",
                "赠送金累计", "赠送金退款", "赠送金抵扣", "赠送金保有",
                "积分累计", "积分消费", "积分保有"
            ]
            ws.append(headers)

            # 写入数据
            for item in all_customer_records:
                row = [
                    item.get("name", ""),
                    item.get("vipNum", ""),
                    item.get("vipLevel", ""),
                    item.get("customerTypeName", ""),
                    item.get("belongCounselorName", ""),
                    item.get("depositAmountTotal", 0),
                    item.get("refundDepositAmount", 0),
                    item.get("deductDepositAmount", 0),
                    item.get("depositAmount", 0),
                    item.get("balanceTotal", 0),
                    item.get("refundBalanceAmount", 0),
                    item.get("deductBalanceAmount", 0),
                    item.get("balance", 0),
                    item.get("givenBalanceTotal", 0),
                    item.get("refundGivenBalanceAmount", 0),
                    item.get("deductGivenBalanceAmount", 0),
                    item.get("givenBalance", 0),
                    item.get("integralTotal", 0),
                    item.get("consumeIntegral", 0),
                    item.get("integral", 0)
                ]
                ws.append(row)

            # 保存到内存中并作为文件返回
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = "客户资产保有统计.xlsx"
            response = FileResponse(buffer, as_attachment=True, filename=filename)
            response["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return response

        # 咨询接诊记录
        elif action == "consultation_admission":
            all_consultation_admission = []
            start_time = request.data.get("startTime", datetime.today().strftime("%Y-%m-%d"))
            end_time = request.data.get("endTime", datetime.today().strftime("%Y-%m-%d"))

            for info in login_infos:
                login_payload = base_login_payload.copy()
                login_payload["tenantId"] = info["tenantId"]

                try:
                    login_resp = requests.post(login_url, json=login_payload, headers={"Content-Type": "application/json"})
                    token = login_resp.json().get("data", {}).get("user", {}).get("token")
                    if not token:
                        continue

                    page = 1
                    size = 1000
                    total_pages = 1

                    while page <= total_pages:
                        payload = {
                            "page": page,
                            "size": size,
                            "types": [],
                            "startTime": start_time,
                            "endTime": end_time,
                            "status": 1
                        }

                        resp = requests.post(
                            admission_url,
                            json=payload,
                            headers={"Content-Type": "application/json", "authorization": token}
                        )

                        data = resp.json().get("data", {})
                        records = data.get("records", [])
                        total_pages = data.get("pages", 1)

                        all_consultation_admission.extend(records)
                        page += 1

                except Exception:
                    continue

            # 写 Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "咨询"

            ws.append(["咨询"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)

            headers = [
                "客户", "电话", "客户类型", "来院类型", "紧急客人", "就诊项目", "咨询项目",
                "接待时间", "已等待", "接诊咨询师", "咨询状态", "状态", "消费状态",
                "表单", "主诉", "治疗方案", "归属咨询师"
            ]
            ws.append(headers)

            for item in all_consultation_admission:
                phone = ""
                phones = item.get("phoneNumber", [])
                if phones and isinstance(phones, list):
                    phone = phones[0].get("phoneNumber", "")

                row = [
                    item.get("customerName", ""),
                    phone,
                    item.get("customerTypeName", ""),
                    item.get("consultationReceptionTypeName", ""),
                    "是" if item.get("urgentFlag") else "否",
                    ", ".join([p.get("label") for p in item.get("pros", [])]) if item.get("pros") else "",
                    item.get("conPros", ""),
                    datetime.fromtimestamp(item.get("receptionTime", 0) / 1000).strftime("%Y-%m-%d %H:%M") if item.get("receptionTime") else "",
                    item.get("waitTime", ""),
                    item.get("userName", ""),
                    item.get("statusName", ""),
                    "作废" if item.get("bargainFlag") is False else "正常",
                    item.get("bargainFlagName", ""),
                    item.get("dynamicFormValue", ""),
                    item.get("customerRemark", ""),
                    item.get("diagnosisLs", ""),
                    item.get("sysUserName", "")
                ]
                ws.append(row)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = "咨询接诊记录.xlsx"
            response = FileResponse(buffer, as_attachment=True, filename=filename)
            response["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return response

